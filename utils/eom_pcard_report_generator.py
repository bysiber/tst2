from datetime import datetime
from enum import Enum
import openpyxl
import pandas as pd
from openpyxl.styles import NamedStyle
from pandas import DataFrame
from config import Paths
from _utils.sharepoint import SharePoint

import config as config
from utils.report_generator_base import ReportGeneratorBase
import utils.emailer as emailer
from utils.xlsx_interactor import ExcelHelper
from _utils.logger import logger
import os
from pathlib import Path


class HotelReportHeaders(Enum):
    full_service = 'Full Service'
    id = 'ID'
    acct = 'ACCT#'
    bank = 'Bank'
    b_u = 'B/U'


class TransactionReportHeader(Enum):
    txn_number = 'Txn Number'
    card_account_number = 'Card Account Number'
    post_date = 'Post Date'
    vendor_name = 'Vendor Name'
    item_total = 'Item Total'
    item_description = 'Item Description'
    department = 'GL: Department'
    department_desc = 'GL: Department Desc'
    general_ledger_account = 'GL: General Ledger Account'
    general_ledger_account_desc = 'GL: General Ledger Account Desc'
    business_unit = 'GL: Business Unit'
    business_unit_desc = 'GL: Business Unit Desc'
    ch_full_name = 'CH Full Name'
    card_embossed_line_2 = 'Card Embossed Line 2'
    grp_full_name = 'Grp Full Name'


class ReportGenerator(ReportGeneratorBase):
    helper: ExcelHelper
    card_numbers_data: DataFrame
    file_filters_data: DataFrame
    payments_data: DataFrame
    pcard_report: DataFrame
    pcard_report_tabs = {}
    fs_report_tabs = {}
    ss_report_tabs = {}
    sh_report_tabs = {}
    rules_filepath_local = Path(os.getcwd(), config.Paths.rules_filepath, config.FileNames.rules_master_filename)
    
    def __init__(self, sharepoint_client: SharePoint = None):
        self.sp = sharepoint_client
        self.dollar_style = 'dollar_style'
        self.amount_style = None

    def initialize_manager(self):
        self.helper = ExcelHelper()

    def read_reports(self):
        #The files are already downloaded
        self.download_master_file()
        
        self.card_numbers_data = self.helper.read_xlsx(config.Paths.rules_filepath,
                                                       config.FileNames.rules_master_filename,
                                                       'SS & FS Cardholders')
        self.file_filters_data = self.helper.read_xlsx(config.Paths.rules_filepath,
                                                       config.FileNames.rules_master_filename,
                                                       '01File Filters')
        self.payments_data = self.helper.read_xlsx(config.Paths.rules_filepath,
                                                   config.FileNames.rules_master_filename,
                                                   'FSPayment File')
        self.pcard_report = self.helper.read_xlsx(config.Paths.downloads_filepath,
                                                  '01 - EOM Pcard Report.xls',
                                                  'sheet 1')

    def complete_card_numbers(self):
        self.helper.complete_8_digits_card_numbers(self.card_numbers_data, self.pcard_report)

    def create_filters(self):
        self.helper.create_file_filters_and_tabs(self.file_filters_data, self.pcard_report, self.pcard_report_tabs)

    def create_specific_files(self):
        self.helper.complete_specific_files(self.pcard_report_tabs, self.fs_report_tabs, self.sh_report_tabs)
        
    def calculate_totals(self):
        tabs_dict_list = [self.pcard_report_tabs, self.fs_report_tabs, self.ss_report_tabs]
        self.helper.complete_totals_from_files(tabs_dict_list)

    def save_new_reports(self):
        reports = config.ReportConfiguration.OUTPUT_REPORTS
        if len(self.sh_report_tabs.items()) == 0:
            reports.remove("Suspended Hotels Report")

        for OUTPUT_REPORT in reports:
            OUTPUT_REPORT = self.helper.remove_extra_spaces(OUTPUT_REPORT)
            report_path = config.Paths.processed_filepath

            with pd.ExcelWriter(f'{report_path}/{OUTPUT_REPORT}.xlsx') as writer:
                for (file, tab), data_frame in self.pcard_report_tabs.items():
                    file = self.helper.remove_extra_spaces(file)
                    if file == OUTPUT_REPORT:
                        data_frame.to_excel(writer, sheet_name=f"{tab}", index=False)
                        
                for (file, tab), data_frame in self.fs_report_tabs.items():
                    file = self.helper.remove_extra_spaces(file)
                    if file == OUTPUT_REPORT:
                        data_frame.to_excel(writer, sheet_name=f"{tab}", index=False)
                        
                for (file, tab), data_frame in self.ss_report_tabs.items():
                    file = self.helper.remove_extra_spaces(file)
                    if file == OUTPUT_REPORT:
                        data_frame.to_excel(writer, sheet_name=f"{tab}", index=False)

                for (file, tab), data_frame in self.sh_report_tabs.items():
                                    file = self.helper.remove_extra_spaces(file)
                                    if file == OUTPUT_REPORT:
                                        data_frame.to_excel(writer, sheet_name=f"{tab}", index=False)
                        
 

    def create_reports(self):
        logger.info("Reading the newly downloaded reports.")
        self.initialize_manager()
        self.read_reports()

        logger.info("Matching all credit card numbers with the rule.")
        self.complete_card_numbers()

        logger.info("Creating all the corresponding tabs for General File: EOM Pcard Report - FS & SS.")
        self.create_filters()

        logger.info("Creating all the corresponding tabs for Specific Files: EOM Full Service and EOM Select Service.")
        self.create_specific_files()
        
        logger.info("Calculating all the total amounts for the reports.")
        self.calculate_totals()

        logger.info("Saving the reports.")
        self.save_new_reports()

        logger.info("Applying styles for reports")
        self.style_reports()

    def style_reports(self):
        transactions_tab_headers = set(h.value for h in TransactionReportHeader)
        hotel_tab_headers = set(h.value for h in HotelReportHeaders)

        for OUTPUT_REPORT in config.ReportConfiguration.OUTPUT_REPORTS:
            OUTPUT_REPORT = self.helper.remove_extra_spaces(OUTPUT_REPORT)
            report_filepath = Path(config.Paths.processed_filepath, OUTPUT_REPORT + '.xlsx')
            
            start_tab_green = 'Colony'
            end_tab_green = 'Corepoint'
            skyblue_color = '#2980B9'
            self.set_tab_colors(report_filepath, start_tab_green, end_tab_green, skyblue_color)

            start_tab_red = 'Corepoint'
            end_tab_red = None
            green_color = '#229954'
            self.set_tab_colors(report_filepath, start_tab_red, end_tab_red, green_color)

            wb = openpyxl.load_workbook(report_filepath)
            for tab in wb.worksheets:
                all_tab_rows = list(tab.iter_rows())
                header_col_ids = {c.value: c.col_idx for c in all_tab_rows[0]}
                current_headers_set = set(header_col_ids.keys())

                if transactions_tab_headers == current_headers_set:
                    # Set width
                    tab.column_dimensions['A'].width = 15
                    tab.column_dimensions['B'].width = 20
                    tab.column_dimensions['C'].width = 20
                    tab.column_dimensions['D'].width = 25
                    tab.column_dimensions['E'].width = 15
                    tab.column_dimensions['F'].width = 35
                    tab.column_dimensions['G'].width = 15
                    tab.column_dimensions['H'].width = 20
                    tab.column_dimensions['I'].width = 25
                    tab.column_dimensions['J'].width = 30
                    tab.column_dimensions['K'].width = 15
                    tab.column_dimensions['L'].width = 35
                    tab.column_dimensions['M'].width = 15
                    tab.column_dimensions['N'].width = 20
                    tab.column_dimensions['O'].width = 40
                    # Set amount format
                    for row in all_tab_rows[1:]:
                        for c in row:
                            if c.col_idx in [header_col_ids[TransactionReportHeader.item_total.value]]:
                                c.style = self._set_amount_style()
                    continue
                if hotel_tab_headers == current_headers_set:
                    # Set width
                    tab.column_dimensions['A'].width = 40
                    tab.column_dimensions['B'].width = 20
                    tab.column_dimensions['C'].width = 20
                    tab.column_dimensions['D'].width = 12
                    tab.column_dimensions['E'].width = 16
                    # Set amount format
                    for row in all_tab_rows[1:]:
                        for c in row:
                            if c.col_idx in [header_col_ids[HotelReportHeaders.b_u.value]]:
                                c.style = self._set_amount_style()
                    continue
                logger.info(f'Could not style the tab: {tab.title}')

            wb.save(report_filepath)
            logger.info(f'Styles were applied for: {OUTPUT_REPORT}')
            
    def set_tab_colors(self, filepath, start_tab, end_tab, color):
        wb = openpyxl.load_workbook(filepath)

        is_green = False

        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]

            if sheetname == start_tab:
                is_green = True
            elif sheetname == end_tab:
                is_green = False

            if is_green:
                sheet.sheet_properties.tabColor = 'FF' + color[1:]

        wb.save(filepath)

    def _set_amount_style(self):
        if not self.amount_style:
            amount_format = r'_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
            amount_style = NamedStyle(name=self.dollar_style)
            amount_style.number_format = amount_format
            self.amount_style = amount_style
        return self.amount_style

    def upload_reports(self, download_directory: str, processed_directory: str):
        sp = self.sp
        if not sp:
            logger.error("Sharepoint connection not detected. Stopping upload.")
            return
        
        report_et_date_converted = datetime.today()
        
        upload_root_path = Path(Paths.sp_processed_filepath, str(report_et_date_converted.year), str(report_et_date_converted.strftime("%m-%d-%Y")))
        upload_root_path_downloads = Path(upload_root_path, "Downloads")
        upload_root_path_processed = Path(upload_root_path, "Processed")
        
        # Uploading processed files.
        processed_files = os.listdir(processed_directory)
        sp.create_folder(upload_root_path_processed)
        for file_name in processed_files:
            file_path = os.path.join(processed_directory, file_name)
            
            if os.path.isfile(file_path):
                processed_report_filepath_local = Path(processed_directory, file_name)
                processed_report_filepath_sp = Path(upload_root_path_processed, file_name)
                
                sp.upload(local_filepath=processed_report_filepath_local, target_filepath=processed_report_filepath_sp)
                logger.info(f'Processed Report uploaded: {file_name}')
            
        # Uploading downloaded files.
        downloaded_files = os.listdir(download_directory)
        sp.create_folder(upload_root_path_downloads)
        for file_name in downloaded_files:
            file_path = os.path.join(download_directory, file_name)

            if os.path.isfile(file_path):
                download_report_filepath_local = Path(download_directory, file_name)
                download_report_filepath_sp = Path(upload_root_path_downloads, file_name)
                
                sp.upload(local_filepath=download_report_filepath_local, target_filepath=download_report_filepath_sp)
                logger.info(f'Downloaded Report uploaded: {file_name}')

        return [upload_root_path_downloads, upload_root_path_processed]

    def download_master_file(self):
        logger.info(f'Downloading Master file')

        sp = self.sp
        if not sp:
            logger.error("Sharepoint connection not detected. Stopping upload.")
            return
        
        sp_master_file_path = Paths.sp_rules_master_filepath
        local_master_file_path = self.rules_filepath_local
        sp.download(relative_filepath=sp_master_file_path, local_filepath=local_master_file_path)
        
    def send_report_emails(self, auxiliary_email: str, processed_directory: str):
        rute = Path(processed_directory)

        if not rute.is_dir():
            logger.error("Email not sent: there are no processed files to attach.")
            return
        
        files = [file for file in rute.glob('**/*') if file.is_file()]
        file_list = list(files)
        emailer.send_email_with_multiple_attachments(auxiliary_email, file_list)

