import locale
import os
from typing import Optional
import openpyxl
import re
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from _utils import config
from logger import logger
import utils.emailer as emailer
from _utils.sharepoint import SharePoint
from utils.report_generator_base import ReportGeneratorBase
from pathlib import Path
from config import Paths, ReportConfiguration, FileNames
import pandas as pd
from enum import Enum
from datetime import datetime
import sys


class ReportHeader(Enum):
    receipt_status = 'Receipt Status'
    gl_acct = 'GL: General Ledger Account'
    ch_sign_of_full_name = 'CH Signoff Full Name'
    p_property = 'Card Embossed Line 2'
    item_description = 'Item Description'
    vpo_vpf = 'VPO/VPF'
    item_total = 'Item Total'
    gl_business_unit = 'GL: Business Unit'
    post_date = 'Post Date'
    # not used in processing:
    txn_number = 'Txn Number'
    vendor_name = 'Vendor Name'
    gl_department = 'GL: Department'
    ch_full_name = 'CH Full Name'
    grp_full_name = 'Grp Full Name'
    mcc_description = 'MCC Description'
    ch_sign_of_date = 'CH Signoff Date'
    mcc = 'MCC'


class Transaction:
    def __init__(self,
                 ch_sign_of_full_name: str,
                 receipt_status: str,
                 gl_acct: str,
                 item_total: float,
                 vpo_vpf: str,
                 p_property: str,
                 gl_business_unit: str,
                 item_description: str,
                 post_date: datetime,
                 txn_number: str,
                 vendor_name: str,
                 gl_department: str,
                 ch_full_name: str,
                 grp_full_name: str,
                 mcc_description: str,
                 ch_sign_of_date: datetime,
                 mcc: str
                 ):
        self.ch_sign_of_full_name = ch_sign_of_full_name
        self.gl_acct = gl_acct
        self.receipt_status = receipt_status
        self.item_total: float = item_total
        self.item_description: str = item_description
        self.property_name: str = p_property
        self.vpo_vpf: str = vpo_vpf
        self.gl_business_unit: str = gl_business_unit
        self.post_date: str = post_date.strftime("%m.%d.%Y")
        self.txn_number = txn_number
        self.vendor_name = vendor_name
        self.gl_department = gl_department
        self.ch_full_name = ch_full_name
        self.grp_full_name = grp_full_name
        self.mcc_description = mcc_description
        self.ch_sign_of_date: str = ch_sign_of_date.strftime("%m.%d.%Y") if ch_sign_of_date else ch_sign_of_date
        self.mcc = mcc

    @property
    def is_sign_of_missing(self):
        return self.ch_sign_of_full_name in (None, '', 'None',)

    @property
    def is_coding_missing(self):
        return self.gl_acct in (None, '', 'None',)

    @property
    def is_receipt_missing(self):
        return self.receipt_status in (None, '', 'None',)


class ReportGenerator(ReportGeneratorBase):
    output_filepath_storage = Path('Reports', 'BIWEEKLY')
    rules_tab_name = 'BI-WEEKLY REPORT RULES'
    rules_filepath_local = Path(os.getcwd(), Paths.rules_filepath, FileNames.rules_master_filename)
    transactions: list

    def __init__(self, sharepoint_client: SharePoint = None):
        self.sp = sharepoint_client
        self.extracted_transactions: Optional[list] = None
        self.raw_input_data: Optional[list] = None
        self.header_ids: Optional[dict] = None
        self.vpo_lookup_rules: Optional[dict] = None
        self.report_period: Optional[str] = None
        self.start_date: Optional[str] = None
        self.end_date: Optional[str] = None
        self.report_et_date: Optional[datetime] = None
        self.dollar_style = 'dollar_style'
        self.amount_style = None
        self.created_reports: Optional[list] = None
        self.lookup_data: Optional[dict] = None

    def read_input_file(self) -> list:
        """
        Convert input file from .xls to .xlsx and extract all raw rows
        if self.raw_input_data is None

        :return: list
        """
        if self.raw_input_data:
            return self.raw_input_data

        initial_input_filename = Path(ReportConfiguration.REPORTS_TO_RUN[1] + '.xls')
        updated_input_filename = Path(ReportConfiguration.REPORTS_TO_RUN[1] + '.xlsx')
        input_filepath_local = Path(os.getcwd(), Paths.downloads_filepath)

        data_frame = pd.read_excel(Path(input_filepath_local, initial_input_filename))
        data_frame.to_excel(Path(input_filepath_local, updated_input_filename), index=False, engine='openpyxl')

        wb = openpyxl.load_workbook(Path(input_filepath_local, updated_input_filename))
        sheet = wb.active

        raw_input_data = list(sheet.iter_rows())
        self.raw_input_data = raw_input_data
        if len(self.raw_input_data) <= 1:
            logger.info('The input report is empty! Finishing the execution')
            sys.exit(0)

        header_row = raw_input_data[0]
        self.header_ids = {h_c.value: header_row.index(h_c) for h_c in header_row}

        all_post_dates = set(row[self.header_ids[ReportHeader.post_date.value]].value for row in self.raw_input_data[1:])
        start_date = min(all_post_dates).strftime("%m.%d")
        end_date = max(all_post_dates).strftime("%m.%d.%y")
        self.report_period = f'{start_date}-{end_date}'
        
        start_date = min(all_post_dates).strftime("%m.%d.%y")
        self.start_date = datetime.strptime(start_date, "%m.%d.%y").strftime("%m.%d.%y")
        self.end_date = datetime.strptime(end_date, "%m.%d.%y").strftime("%m.%d.%y")
        self.report_et_date = end_date

        return self.raw_input_data

    def read_vpo_lookup_rules(self) -> dict:
        """
        Rules structure {"VPO ASSIGNMENT": [("Property", "Business Unit"),],}
        :return: dict of rules
        """
        if self.vpo_lookup_rules:
            return self.vpo_lookup_rules
        
        # Download rules file
        self.download_master_file()

        wb = openpyxl.load_workbook(self.rules_filepath_local)
        sheet = wb[self.rules_tab_name]
        all_rows = list(sheet.iter_rows(min_row=2))

        column_mapping = {}
        property_indices = []
        cc_email_indices = []
        vpo_assignment_indices = []
        to_email_indices = []
        
        for cell in sheet[1]:
            column_name = cell.value
            if column_name:
                if column_name.startswith("Property"):
                    property_indices.append(cell.column - 1)
                elif column_name.startswith("CC: Email"):
                    cc_email_indices.append(cell.column - 1)
                elif column_name.startswith("TO: Email"):
                    to_email_indices.append(cell.column - 1)
                elif column_name.startswith("VPO"):
                    vpo_assignment_indices.append(cell.column - 1)
                elif column_name in column_mapping:
                    column_mapping[column_name].append(cell.column - 1)
                else:
                    column_mapping[column_name] = [cell.column - 1]
                    
        column_mapping['Property (SS & FS)'] = property_indices
        column_mapping['CC: Email Addresses'] = cc_email_indices
        column_mapping['VPO Assignment'] = vpo_assignment_indices
        column_mapping['TO: Email Addresses'] = to_email_indices
        vpo_rules = {}
        lookup_data_list: Optional[list[dict]] = []
    
        # Testing purposes, the one commented validates that it corresponds to a HighGate's email
        # email_pattern = r"\b[A-Za-z0-9._%+-]+@highgate\.com\b"
        email_pattern = r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b"
        
        for row_number, row in enumerate(all_rows, start=2):
            vpo = str(row[column_mapping["VPO Assignment"][-1]].value)
            property_name = str(row[column_mapping["Property (SS & FS)"][-1]].value)
            if not all([vpo,
                        property_name]):
                continue

            if vpo_rules.get(vpo, False):
                vpo_rules[vpo].append(property_name)
            else:
                vpo_rules[vpo] = [property_name]
                
            vpo_assignment = str(row[column_mapping["VPO Assignment"][-1]].value)
            to_email_addresses = str(row[column_mapping["TO: Email Addresses"][-1]].value)
            cc_email_addresses = []
            cc_column_indices = [column_mapping[key] for key in column_mapping.keys() if key.startswith("CC: Email Addresses")]
            cc_column_indices = [item for sublist in cc_column_indices for item in sublist]

            for index in cc_column_indices:
                cell_value = str(row[index].value)
                
                if re.match(email_pattern, cell_value):
                    cc_email_addresses.append(cell_value)
            
            if not (vpo or to_email_addresses):
                logger.warning(f"The Row number: {row_number} does not have any vpo or email specified. Continuing with the process.")
                continue
            
            lookup_data = {
                "VPO": vpo_assignment,
                "recipients": [to_email_addresses],
                "recipientsCC": cc_email_addresses
            }
            lookup_data_list.append(lookup_data)
            
        # Creating a set based on unique values from VPOs, and agruping recipientsCC corresponding to those VPOs.
        vpo_recipientsCC = {}
        for _dict in lookup_data_list:
            vpo = _dict['VPO']
            recipients = _dict.get('recipients', [])
            recipientsCC = _dict['recipientsCC']
            if vpo in vpo_recipientsCC:
                vpo_recipientsCC[vpo]['recipientsCC'].extend(recipientsCC)
            else:
                vpo_recipientsCC[vpo] = {'VPO': vpo, 'recipients': recipients, 'recipientsCC': recipientsCC}
        for vpo_info in vpo_recipientsCC.values():
            vpo_info['recipientsCC'] = list(set(vpo_info['recipientsCC']))

        self.lookup_data = list(vpo_recipientsCC.values())
            
        self.vpo_lookup_rules = vpo_rules
        return self.vpo_lookup_rules

    def extract_transactions(self) -> list[Transaction]:
        """
        Extracts transactions only with defined VPO via rules
        :return:
        """
        if self.extracted_transactions:
            return self.extracted_transactions

        extracted_transactions = []
        # For each raw data row
        for raw_data in self.raw_input_data[1:]:
            transaction_created = False
            # Try to apply rule from each VPO
            for vpo_name, vpo_data in self.vpo_lookup_rules.items():
                for property_name in vpo_data:
                    # Create transaction if matched
                    row_property_name = raw_data[self.header_ids[ReportHeader.p_property.value]].value
                    if property_name.casefold() == row_property_name.casefold():
                        extracted_transactions.append(
                            Transaction(
                                ch_sign_of_full_name=raw_data[self.header_ids[ReportHeader.ch_sign_of_full_name.value]].value,
                                receipt_status=raw_data[self.header_ids[ReportHeader.receipt_status.value]].value,
                                gl_acct=raw_data[self.header_ids[ReportHeader.gl_acct.value]].value,
                                item_total=float(raw_data[self.header_ids[ReportHeader.item_total.value]].value),
                                vpo_vpf=vpo_name,
                                p_property=row_property_name,
                                item_description=str(raw_data[self.header_ids[ReportHeader.item_description.value]].value),
                                gl_business_unit=str(raw_data[self.header_ids[ReportHeader.gl_business_unit.value]].value),
                                post_date=raw_data[self.header_ids[ReportHeader.post_date.value]].value,
                                vendor_name=str(
                                    raw_data[self.header_ids[ReportHeader.vendor_name.value]].value),
                                gl_department=raw_data[self.header_ids[ReportHeader.gl_department.value]].value,
                                ch_full_name=str(
                                    raw_data[self.header_ids[ReportHeader.ch_full_name.value]].value),
                                ch_sign_of_date=raw_data[self.header_ids[ReportHeader.ch_sign_of_date.value]].value,
                                grp_full_name=str(
                                    raw_data[self.header_ids[ReportHeader.grp_full_name.value]].value),
                                mcc_description=str(
                                    raw_data[self.header_ids[ReportHeader.mcc_description.value]].value),
                                mcc=str(
                                    raw_data[self.header_ids[ReportHeader.mcc.value]].value),
                                txn_number=str(
                                    raw_data[self.header_ids[ReportHeader.txn_number.value]].value),
                            )
                        )
                        transaction_created = True
                        break
                if transaction_created:
                    break

        self.extracted_transactions = extracted_transactions
        return self.extracted_transactions

    def generate_summary_tab(self, wb: Workbook, vpo_transactions: list[Transaction]) -> Worksheet:
        sheet = wb.create_sheet('Summary', 0)
        # Generate Table 1 data: Hotel | Total within the Grand Total
        hotel_totals = {}
        for tr in vpo_transactions:
            if hotel_totals.get(tr.property_name, False):
                hotel_totals[tr.property_name] = round(hotel_totals[tr.property_name] + tr.item_total, 2)
            else:
                hotel_totals[tr.property_name] = round(tr.item_total, 2)
        grand_total = round(sum(list(hotel_totals.values())), 2)
        # Generate Table 2 data: Transactions w/o receipts | Missing Coding | Missing sign off - as counts
        missing_receipts_count = sum([1 for tr in vpo_transactions if tr.is_receipt_missing])
        missing_coding_count = sum([1 for tr in vpo_transactions if tr.is_coding_missing])
        missing_sign_of_count = sum([1 for tr in vpo_transactions if tr.is_sign_of_missing])
        # Generate Table 3 data: Owner, Hotel, Item Description, Item Total within sub-totals and the Grand Total
        hotel_transactions: dict = {}  # {'Hotel 1': [(Transaction.description, Transaction.amount,)], }
        for tr in vpo_transactions:
            if hotel_transactions.get(tr.property_name, False):
                hotel_transactions[tr.property_name].append((tr.item_description, tr.item_total,))
            else:
                hotel_transactions[tr.property_name] = [(tr.item_description, tr.item_total,)]
        # Add totals per hotel
        for hotel, transactions in hotel_transactions.items():
            hotel_total = round(sum([t[1] for t in transactions]), 2)

            transactions_d = {}
            for tr in transactions:
                if transactions_d.get(tr[0], False):
                    transactions_d[tr[0]] += tr[1]
                else:
                    transactions_d[tr[0]] = tr[1]
            transactions = [(td, round(ta, 2),) for td, ta in transactions_d.items()]
            hotel_transactions[hotel] = (transactions, hotel_total,)

        header_1 = ['Hotel', 'Total', None, 'Hotel', 'Total']
        # Add the first head for Tables 1, 3
        sheet.append(header_1)
        for h, a in hotel_totals.items():
            # Add hotel , total amount | Table 1
            sheet.append([h, a])
        # Add grand total row
        sheet.append(['Grand Total', grand_total])
        # Add table 2 after blank row
        sheet.append([None])
        sheet.append(['Transactions w\o receipts', missing_receipts_count])
        sheet.append(['Missing Coding', missing_coding_count])
        sheet.append(['Missing sign off', missing_sign_of_count])
        # Add Table 3 data on the right after Table 1 after 1 blank column
        start_row = 2
        start_coll = 4
        for hotel, data in hotel_transactions.items():
            transactions = data[0]
            total = data[1]
            row_values = [hotel, total]
            # Add Hotel Sub Header
            for v in row_values:
                sub_header_cell = sheet.cell(row=start_row, column=start_coll)
                sub_header_cell.value = v
                start_coll += 1
            start_row += 1
            start_coll = 4
            # Add Hotel transactions
            for tr in transactions:
                for v in tr:
                    cell = sheet.cell(row=start_row, column=start_coll)
                    cell.value = v
                    start_coll += 1
                start_row += 1
                start_coll = 4
        grand_total_values = ['Grand Total', grand_total]
        # Add Grand Total for Table 3
        for v in grand_total_values:
            cell = sheet.cell(row=start_row, column=start_coll)
            cell.value = v
            start_coll += 1
        start_row += 1
        return sheet

    def generate_transactions_tab(self, wb: Workbook, vpo_transactions: list[Transaction]) -> Worksheet:
        sheet = wb.active
        sheet.title = f'{self.report_period} All Transactions'
        header_row = sorted(self.header_ids, key=self.header_ids.get)
        # Add header row
        sheet.append(header_row)
        # Add all vpo transactions data
        for tr in vpo_transactions:
            sheet.append([
                tr.txn_number,
                tr.post_date,
                tr.vendor_name,
                tr.item_total,
                tr.item_description,
                tr.gl_department,
                tr.gl_acct,
                tr.gl_business_unit,
                tr.ch_full_name,
                tr.property_name,
                tr.grp_full_name,
                tr.mcc_description,
                tr.receipt_status,
                tr.ch_sign_of_date,
                tr.ch_sign_of_full_name,
                tr.mcc
            ])

        return sheet

    def _set_amount_style(self):
        if not self.amount_style:
            amount_format = r'_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
            amount_style = NamedStyle(name=self.dollar_style)
            amount_style.number_format = amount_format
            self.amount_style = amount_style
        return self.amount_style

    def apply_transactions_tab_styles(self, tab: Worksheet):
        logger.info(f'Applying Transactions tab styles')
        all_rows = list(tab.iter_rows())
        
        # Header to bold
        header_row = all_rows[0]
        for c in header_row:
            c.font = Font(bold=True)
            
        # Column Width
        tab.column_dimensions['A'].width = 15
        tab.column_dimensions['B'].width = 15
        tab.column_dimensions['C'].width = 30
        tab.column_dimensions['D'].width = 15
        tab.column_dimensions['E'].width = 40
        tab.column_dimensions['F'].width = 15
        tab.column_dimensions['G'].width = 15
        tab.column_dimensions['H'].width = 15
        tab.column_dimensions['I'].width = 25
        tab.column_dimensions['J'].width = 30
        tab.column_dimensions['K'].width = 35
        tab.column_dimensions['L'].width = 35
        tab.column_dimensions['M'].width = 15
        tab.column_dimensions['N'].width = 15
        tab.column_dimensions['O'].width = 25
        tab.column_dimensions['P'].width = 10
        
        # DataTypes
        amount_cells = [r[3] for r in all_rows[1:]]
        for c in amount_cells:
            c.style = self._set_amount_style()
            
        # Freeze panes
        tab.freeze_panes = "A2"
        
        # Check for empty cells and apply pink color
        pink_fill = PatternFill(start_color="FBFBD7", end_color="FBFBD7", fill_type="solid")
        for row in all_rows[1:]:
            # Check if any of the specified columns are empty
            # ("GL: General Ledger Account", "Receipt Status", "CH Signoff Date")
            if not all([row[col_index].value for col_index in [6, 12, 13]]):  
                for cell in row:
                    cell.fill = pink_fill

        # Hide 3 columns requested by the client
        columns_to_hide = ["Grp Full Name", "MCC Description", "MCC"]
        for col in range(1, tab.max_column + 1):
            column_letter = openpyxl.utils.get_column_letter(col)
            if tab[column_letter + '1'].value in columns_to_hide:
                tab.column_dimensions[column_letter].hidden = True

        

    def apply_summary_tab_styles(self, tab: Worksheet):
        logger.info(f'Applying Summary tab styles')
        all_rows = list(tab.iter_rows())
        # Header to bold
        header_row = all_rows[0]
        for c in header_row:
            c.font = Font(bold=True)
        # Width
        tab.column_dimensions['A'].width = 35
        tab.column_dimensions['B'].width = 15
        tab.column_dimensions['D'].width = 45
        tab.column_dimensions['E'].width = 15
        # DataTypes
        for r in all_rows[1:]:
            r[1].style = self._set_amount_style()
            r[4].style = self._set_amount_style()
        # Freeze panes
        tab.freeze_panes = "A2"
        # Get Hotel Names
        hotels = []
        grand_total_1_row = []
        for r in all_rows[1:]:
            hotel = r[0].value
            if hotel == 'Grand Total':
                grand_total_1_row = r
                break
            hotels.append(hotel)
        border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # Set Bold and Border to Grand Total cells
        for c in grand_total_1_row[:2]:
            c.font = Font(bold=True)
            c.border = border
        # Reset the dollar format for counters (Table 3)
        for i in range(1, 4):
            table3_title_cell = tab.cell(row=grand_total_1_row[0].row + i+1, column=1)
            table3_count_cell = tab.cell(row=grand_total_1_row[0].row + i+1, column=2)
            table3_count_cell.number_format = '#,##0'

            table3_title_cell.font = Font(bold=True)
            table3_count_cell.font = Font(bold=True)
            table3_title_cell.border = border
            table3_count_cell.border = border
        # Set border and bold to Table 2
        for r in all_rows[1:]:
            if r[3].value in hotels or r[3].value == 'Grand Total':
                hotel_title_cell = r[3]
                total_cell = r[4]
                hotel_title_cell.font = Font(bold=True)
                total_cell.font = Font(bold=True)
                hotel_title_cell.border = border
                total_cell.border = border
            if r[3].value == 'Grand Total':
                break

    def sort_tab_by_column(self, tab: Worksheet, column_name):
        """
        Sorts the worksheet by the specified column.

        Args:
        - tab: The worksheet to be sorted.
        - column_name: The name of the column to sort by.

        Returns:
        - None
        """
        # Get the index of the specified column
        column_index = None
        for col in tab.iter_cols(min_row=1, max_row=1):
            if col[0].value == column_name:
                column_index = col[0].column
                break

        if column_index is None:
            print(f"Column '{column_name}' not found.")
            return

        # Get data from the worksheet without the header row
        data = [row for row in tab.iter_rows(min_row=2, values_only=True)]
        # Sort the data by the specified column
        sorted_data = sorted(data, key=lambda x: locale.strxfrm(str(x[column_index - 1])))
        # Clear all rows in the worksheet
        tab.delete_rows(2, tab.max_row)
        # Write the sorted data back to the worksheet
        for row in sorted_data:
            tab.append(row)


    def create_reports(self, processed_directory: str) -> list[Path]:
        # Read Input file
        logger.info("Reading input files...")
        self.read_input_file()

        logger.info("Reading VPO lookup files...")
        # Read VPO rules
        self.read_vpo_lookup_rules()

        logger.info("Extracting transactions from raw data...")
        # Extract Transaction from raw data
        self.extract_transactions()

        created_reports = []
        for vpo in self.lookup_data:
            logger.info(f'Creating report for: {vpo["VPO"]}')
            vpo_transactions = [tr for tr in self.extracted_transactions if tr.vpo_vpf == vpo["VPO"]]
            report_name = f'{self.report_period} BI-WEEKLY {vpo["VPO"]}.xlsx'
            report_filepath = Path(Paths.processed_filepath, report_name)
            wb = openpyxl.Workbook()
            all_transactions_tab = self.generate_transactions_tab(wb, vpo_transactions)
            summary_tab = self.generate_summary_tab(wb, vpo_transactions)
            
            emailer.generate_paths(vpo, processed_directory, report_name)
            
            # Set styles, filters and ordering on tabs
            self.sort_tab_by_column(all_transactions_tab, "Card Embossed Line 2")
            self.apply_transactions_tab_styles(all_transactions_tab)
            self.apply_summary_tab_styles(summary_tab)
            wb.save(report_filepath)
            
            created_reports.append(report_filepath)
            logger.info(f'Report for: {vpo["VPO"]} successfully created')

        self.created_reports = created_reports
        return created_reports

    def upload_reports(self, download_directory: str, processed_directory: str):
        sp = self.sp
        if not sp:
            logger.error("Sharepoint connection not detected. Stopping upload.")
            return
        
        report_et_date_converted = datetime.today()

        upload_root_path = Path(Paths.sp_processed_filepath, str(report_et_date_converted.year), str(report_et_date_converted.strftime("%m-%d-%Y")))
        upload_root_path_downloads = Path(upload_root_path, "Downloads")
        upload_root_path_processed = Path(upload_root_path, "Processed")
        
        # Uploading processed files.
        sp.create_folder(upload_root_path_processed)
        for report_filepath_local in self.created_reports:
            report_filepath_sp = Path(upload_root_path_processed, report_filepath_local.name)
            sp.upload(local_filepath=report_filepath_local, target_filepath=report_filepath_sp)
            logger.info(f'Processed Report uploaded: {report_filepath_local.name}')
            
        # Uploading downloaded files.
        files = os.listdir(download_directory)
        sp.create_folder(upload_root_path_downloads)
        for file_name in files:
            file_path = os.path.join(download_directory, file_name)

            if os.path.isfile(file_path):
                download_report_filepath_local = Path(download_directory, file_name)
                download_report_filepath_sp = Path(upload_root_path_downloads, file_name)
                
                sp.upload(local_filepath=download_report_filepath_local, target_filepath=download_report_filepath_sp)
                logger.info(f'Downloaded Report uploaded: {file_name}')

        return [upload_root_path_downloads, upload_root_path_processed]

    def download_master_file(self):
        logger.info(f'Downloading Master file')
        
        sp = self.sp
        if not sp:
            return
        sp_master_file_path = Paths.sp_rules_master_filepath
        local_master_file_path = self.rules_filepath_local
        sp.download(relative_filepath=sp_master_file_path, local_filepath=local_master_file_path)

    def send_report_emails(self, auxiliary_email: str, processed_directory: str):
        rute = Path(processed_directory)
        if not rute.is_dir():
            logger.error("Emails not sent: there are no processed files to attach.")
            raise Exception("Emails not sent: there are no processed files to attach.")
        
        self.lookup_data = self.validate_files_not_empty(self.lookup_data)
        
        if not auxiliary_email:
            logger.info(f"Lookup data: {self.lookup_data}")
            emailer.send_report_emails(self.lookup_data, self.start_date, self.end_date)
        else:
            logger.info(f"Auxiliary email was specified. Now triggering emails to the following email box: {auxiliary_email}")
            emailer.send_auxiliary_email(self.lookup_data, auxiliary_email, self.start_date, self.end_date)
        return
    
    def validate_files_not_empty(self, lookup_data: list):
        if lookup_data is None or len(lookup_data) == 0:
            logger.error("Emails not sent: the lookup data structure was not created correctly.")
            raise Exception("Emails not sent: the lookup data structure was not created correctly.")
        
        new_lookup_data = []
        
        for data in lookup_data:
            file_path = data["fpaths"]
            file = Path(file_path)
            
            if file.suffix == '.xlsx':
                wb = openpyxl.load_workbook(file)
                all_transactions_sheet = next((wb[sheet] for sheet in wb.sheetnames if 'All Transactions' in sheet), None)
            
                if all_transactions_sheet is None:
                    logger.error(f"Emails not sent: The file ({file.name}) does not contain a tab named 'All Transactions'.")
                    raise ValueError(f"Emails not sent: The file ({file.name}) does not contain a tab named 'All Transactions'.")
                
                if all_transactions_sheet.max_row <= 1 and all_transactions_sheet.cell(2, 2).value is None:
                    logger.warning(f"The tab 'All Transactions' from file ({file.name}) is empty and will be excluded from further processing.")
                else:
                    new_lookup_data.append(data)
        
        return new_lookup_data
                
   

